import os
import re
import pytesseract
from pdf2image import convert_from_path
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

excel_file = "MacroBilanci.xlsx"
sheet_name = "Risultati"
pdf_folder = "Bilanci"

pdf_folder_path = os.path.join(os.getcwd(), pdf_folder)

pattern = r"(somministrati|somministrazione|interinali|agency workers|staff leasing workers|temporary workers)[^\d]{0,50}(\d{1,6})"

def extract_text_from_pdf(pdf_path):
    images = convert_from_path(pdf_path, dpi=300)
    text = ""
    for img in images:
        text += pytesseract.image_to_string(img, lang="ita+eng") + "\n"
    return text

def extract_year(filename):
    match = re.search(r"\d{4}", filename)
    return match.group(0) if match else None

def extract_value(text):
    matches = re.findall(pattern, text, re.IGNORECASE)
    if matches:
        numbers = [int(m[1]) for m in matches]
        return max(numbers)
    return None

results = []
pdf_files = [f for f in os.listdir(pdf_folder_path) if f.lower().endswith(".pdf")]

for file in tqdm(pdf_files, desc="Analisi PDF", unit="file"):
    pdf_path = os.path.join(pdf_folder_path, file)
    text = extract_text_from_pdf(pdf_path)
    anno = extract_year(file)
    valore = extract_value(text)
    results.append({"File/Azienda": file, "Anno": anno, "Valore somministrati": valore})

df = pd.DataFrame(results)
wb = load_workbook(excel_file)
ws = wb[sheet_name]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

for i, row in df.iterrows():
    ws.cell(row=i+2, column=1, value=row["File/Azienda"])
    ws.cell(row=i+2, column=2, value=row["Anno"])
    ws.cell(row=i+2, column=3, value=row["Valore somministrati"])

wb.save(excel_file)
print("✅ Aggiornamento completato!")
